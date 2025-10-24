import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from django.core.files import File
from apps.rosstat_parser.models import Region
from apps.rosstat_parser.api.v1.services.table_commands import (get_last_filled_row, write_months_in_row,
                                                                find_cell_by_value, get_district_and_regions_by_region,
                                                                get_last_filled_column, get_row_average)


os.makedirs("tables", exist_ok=True)

def create_an_turpotok_table(region):
    turpotok_table = load_workbook("downloads/Turpotok_08-2025.xlsx")
    analytical_table = Workbook()
    turpotok = analytical_table.active
    turpotok.title = "Турпоток"

    turpotok["A1"] = "Поездки"
    turpotok["A3"] = "Годовые данные"
    turpotok["A5"] = "Год"
    turpotok["B5"] = "Значение"
    turpotok["C5"] = "Динамика"

    for year in range(2022, datetime.now().year + 1):
        sheet_name = f"1.1.{year}"
        if sheet_name not in turpotok_table.sheetnames:
            break

        turpotok_sheet_11 = turpotok_table[sheet_name]
        region_row = find_cell_by_value(turpotok_sheet_11, region)[0]
        row_index = get_last_filled_row(turpotok, "A") + 1
        turpotok[f"A{row_index}"] = year
        turpotok[f"B{row_index}"] = turpotok_sheet_11[f"B{region_row}"].value
        turpotok[f"C{row_index}"] = f"=ROUND((B{row_index}/B{row_index - 1})*100, 0)%"

    current_row = get_last_filled_row(turpotok, "A") + 3
    turpotok[f"A{current_row}"] = "Месячные\nОперативные\nДанные"
    turpotok[f"A{current_row + 1}"] = "Период"
    write_months_in_row(turpotok, f"B{current_row + 1}")

    for year in range(2022, datetime.now().year + 1):
        sheet_name = f"3.{year}"
        if sheet_name not in turpotok_table.sheetnames:
            break

        turpotok_sheet_3 = turpotok_table[sheet_name]
        region_row = find_cell_by_value(turpotok_sheet_3, region)[0]
        row_index = get_last_filled_row(turpotok, "A") + 1
        turpotok[f"A{row_index}"] = f"{year},\nнарастающим\nитогом"
        turpotok[f"A{row_index + 1}"] = f"{year}, помесячно"
        turpotok[f"A{row_index + 2}"] = f"{year}, доля\n месяца"

        for i in range(1, 13):
            turpotok.cell(row=row_index, column=i + 1, value=turpotok_sheet_3.cell(row=region_row, column=i + 1).value)

            if i != 1:
                turpotok.cell(
                    row=row_index + 1, column=i + 1,
                    value=f"={turpotok.cell(row=row_index, column=i + 1).value}-{turpotok.cell(row=row_index, column=i).value}"
                )
            else:
                turpotok.cell(row=row_index + 1, column=i + 1, value=turpotok_sheet_3.cell(row=region_row, column=i + 1).value)

            turpotok.cell(
                row=row_index + 2, column=i + 1,
                value=f"=ROUND(({str(turpotok.cell(row=row_index + 1, column=i + 1).value)[1::]}) / {turpotok_sheet_3[f'M{region_row}'].value}*100, 0)%"
            )

    current_row = get_last_filled_row(turpotok, "A") + 3
    turpotok[f"A{current_row}"] = f"Динамика роста месяц к месяцу"
    turpotok[f"A{current_row + 1}"] = "Период"
    write_months_in_row(turpotok, f"B{current_row + 1}")

    for year in range(2022, datetime.now().year):
        row_index = get_last_filled_row(turpotok, "A") + 1
        turpotok[f"A{row_index}"] = f"{year + 1}/{year}"

        for i in range(2, 14):
            first_year_value_coord = find_cell_by_value(turpotok, f"{year}, помесячно")
            if i != 2:
                first_year_value = str(turpotok.cell(
                    row=first_year_value_coord[0], column=i).value
                                       )[1:]

                second_year_value_coord = find_cell_by_value(turpotok, f"{year + 1}, помесячно")
                second_year_value = str(turpotok.cell(
                    row=second_year_value_coord[0], column=i).value
                                        )[1:]
            else:
                first_year_value = str(turpotok.cell(
                    row=first_year_value_coord[0], column=i).value
                                       )

                second_year_value_coord = find_cell_by_value(turpotok, f"{year + 1}, помесячно")
                second_year_value = str(turpotok.cell(
                    row=second_year_value_coord[0], column=i).value
                                        )

            turpotok.cell(
                row=row_index, column=i,
                value=f"=ROUND(({second_year_value})/({first_year_value}) * 100, 0)%"
            )

    current_row = get_last_filled_row(turpotok, "A") + 3
    turpotok[f"A{current_row}"] = f"Ночевки"
    turpotok[f"A{current_row + 1}"] = "Год"
    turpotok[f"B{current_row + 1}"] = "Значение"
    turpotok[f"C{current_row + 1}"] = "Динамика"

    for year in range(2022, datetime.now().year + 1):
        sheet_name = f"1.2.{year}"
        if sheet_name not in turpotok_table.sheetnames:
            break

        turpotok_sheet_12 = turpotok_table[sheet_name]
        region_row = find_cell_by_value(turpotok_sheet_12, region)[0]
        row_index = get_last_filled_row(turpotok, "A") + 1
        turpotok[f"A{row_index}"] = year
        turpotok[f"B{row_index}"] = turpotok_sheet_12[f"B{region_row}"].value
        turpotok[f"C{row_index}"] = f"=ROUND((B{row_index}/B{row_index - 1})*100, 0)%"

    current_row = get_last_filled_row(turpotok, "A") + 3
    turpotok[f"A{current_row}"] = f"Продолжительность пребывания"
    turpotok[f"A{current_row + 1}"] = "Год"
    turpotok[f"B{current_row + 1}"] = "Значение"
    turpotok[f"C{current_row + 1}"] = "Динамика"

    for year in range(2022, datetime.now().year + 1):
        row_index = get_last_filled_row(turpotok, "A") + 1
        turpotok[f"A{row_index}"] = year

        nights_value = find_cell_by_value(turpotok, 'Ночевки')
        trips_value = find_cell_by_value(turpotok, 'Поездки')
        try:
            turpotok[f"B{row_index}"] = round((int(turpotok.cell(row=nights_value[0] + (year - 2020), column=2).value) /
                                         int(turpotok.cell(row=trips_value[0] + (year - 2017), column=2).value)), 1)
        except: continue

        turpotok[f"C{row_index}"] = f'=ROUND(B{row_index}/B{row_index - 1}*100, 2)&"%"'

    analytical_table.save(f"tables/Аналитические таблицы {region}.xlsx")


def create_an_mesto_no_table(region):
    turpotok_table = load_workbook("downloads/Turpotok_08-2025.xlsx")
    analytical_table = load_workbook(f"tables/Аналитические таблицы {region}.xlsx")
    sheet_name = "Место НО"
    if sheet_name in analytical_table.sheetnames:
        mesto_sheet = analytical_table[sheet_name]
    else:
        mesto_sheet = analytical_table.create_sheet(title=sheet_name)

    mesto_sheet["A1"] = "Поездки"
    mesto_sheet["A3"] = "Регион"
    mesto_sheet["A4"] = "Российская Федерация"

    district, regions = get_district_and_regions_by_region(
        path="downloads/Turpotok_08-2025.xlsx",
        sheet_name="1.1.2022",
        target_region=region
    )

    mesto_sheet["A5"] = district

    for year in range(2022, datetime.now().year + 1):
        sheet_name = f"1.1.{year}"
        if sheet_name not in turpotok_table.sheetnames:
            break

        turpotok_sheet_11 = turpotok_table[sheet_name]
        current_col = get_last_filled_column(mesto_sheet, 3) + 1

        mesto_sheet.cell(row=3, column=current_col, value=year)
        mesto_sheet.cell(row=3, column=current_col + 1, value=f"{year}, Доля в РФ")
        mesto_sheet.cell(row=3, column=current_col + 2, value=f"{year}, Доля в ПФО")

        mesto_sheet.cell(row=4, column=current_col, value=turpotok_sheet_11["B7"].value)
        okrug_coord = find_cell_by_value(turpotok_sheet_11, district)

        mesto_sheet.cell(row=5, column=current_col, value=turpotok_sheet_11.cell(
            row=okrug_coord[0], column=okrug_coord[1] + 1
        ).value)

        mesto_sheet.cell(
            row=5, column=current_col + 1,
            value=f"=ROUND(({mesto_sheet.cell(row=5, column=current_col).value} / "
                  f"{mesto_sheet.cell(row=4, column=current_col).value}) * 100, 0)%"
        )

    for reg in regions:
        index = regions.index(reg)
        mesto_sheet[f"A{index + 6}"] = reg

        for year in range(2022, datetime.now().year + 1):
            sheet_name = f"1.1.{year}"
            if sheet_name not in turpotok_table.sheetnames:
                break

            turpotok_sheet_11 = turpotok_table[sheet_name]

            region_coord = find_cell_by_value(turpotok_sheet_11, reg)
            current_col = find_cell_by_value(mesto_sheet, f"{year}, Доля в РФ")[1] - 1

            mesto_sheet.cell(
                row=index + 6, column=current_col,
                value=turpotok_sheet_11.cell(row=region_coord[0], column=region_coord[1] + 1).value
            )

            mesto_sheet.cell(
                row=index + 6, column=current_col + 1,
                value=f"=ROUND(({mesto_sheet.cell(row=index + 6, column=current_col).value} / "
                      f"{mesto_sheet.cell(row=4, column=current_col).value}) * 100, 2)%"
            )

            mesto_sheet.cell(
                row=index + 6, column=current_col + 2,
                value=f"=ROUND(({mesto_sheet.cell(row=index + 6, column=current_col).value} / "
                      f"{mesto_sheet.cell(row=5, column=current_col).value}) * 100, 2)%"
            )

    current_row = get_last_filled_row(mesto_sheet, "A") + 3
    mesto_sheet[f"A{current_row}"] = f"Ночевки"
    mesto_sheet[f"A{current_row + 2}"] = "Регион"
    mesto_sheet[f"A{current_row + 3}"] = "Российская Федерация"

    district, regions = get_district_and_regions_by_region(
        path="downloads/Turpotok_08-2025.xlsx",
        sheet_name="1.2.2022",
        target_region=region
    )

    mesto_sheet[f"A{current_row + 4}"] = district

    for year in range(2022, datetime.now().year + 1):
        sheet_name = f"1.2.{year}"
        if sheet_name not in turpotok_table.sheetnames:
            break

        turpotok_sheet_12 = turpotok_table[sheet_name]
        current_col = get_last_filled_column(mesto_sheet, current_row + 2) + 1

        mesto_sheet.cell(row=current_row + 2, column=current_col, value=year)
        mesto_sheet.cell(row=current_row + 2, column=current_col + 1, value=f"{year}, Доля в РФ")
        mesto_sheet.cell(row=current_row + 2, column=current_col + 2, value=f"{year}, Доля в ПФО")

        mesto_sheet.cell(row=current_row + 3, column=current_col, value=turpotok_sheet_12["B7"].value)
        okrug_coord = find_cell_by_value(turpotok_sheet_12, district)

        mesto_sheet.cell(row=current_row + 4, column=current_col, value=turpotok_sheet_12.cell(
            row=okrug_coord[0], column=okrug_coord[1] + 1
        ).value)

        mesto_sheet.cell(
            row=current_row + 4, column=current_col + 1,
            value=f"=ROUND(({mesto_sheet.cell(row=current_row + 4, column=current_col).value} / "
                  f"{mesto_sheet.cell(row=current_row + 3, column=current_col).value}) * 100, 0)%"
        )

    for reg in regions:
        index = regions.index(reg) + current_row + 5
        mesto_sheet[f"A{index}"] = reg

        for year in range(2022, datetime.now().year + 1):
            sheet_name = f"1.2.{year}"
            if sheet_name not in turpotok_table.sheetnames:
                break

            turpotok_sheet_12 = turpotok_table[sheet_name]

            region_coord = find_cell_by_value(turpotok_sheet_12, reg)
            current_col = find_cell_by_value(mesto_sheet, f"{year}, Доля в РФ")[1] - 1

            mesto_sheet.cell(
                row=index, column=current_col,
                value=turpotok_sheet_12.cell(row=region_coord[0], column=region_coord[1] + 1).value
            )

            mesto_sheet.cell(
                row=index, column=current_col + 1,
                value=f'=ROUND({mesto_sheet.cell(row=index, column=current_col).value} / '
                      f'{mesto_sheet.cell(row=current_row + 3, column=current_col).value} * 100, 2)&"%"'
            )

            mesto_sheet.cell(
                row=index, column=current_col + 2,
                value=f'=ROUND(({mesto_sheet.cell(row=index, column=current_col).value} / '
                      f'{mesto_sheet.cell(row=current_row + 4, column=current_col).value}) * 100, 2)&"%"'
            )

    analytical_table.save(f"tables/Аналитические таблицы {region}.xlsx")


def create_an_ksr_table(region):
    ksr_year_table = load_workbook("downloads/KSR_god_sub_2024.xlsx")
    ksr_month_table = load_workbook("downloads/KSR_mes_sub_08-2025.xlsx")
    analytical_table = load_workbook(f"tables/Аналитические таблицы {region}.xlsx")
    sheet_name = "КСР"
    if sheet_name in analytical_table.sheetnames:
        ksr_sheet = analytical_table[sheet_name]
    else:
        ksr_sheet = analytical_table.create_sheet(title=sheet_name)

    ksr_sheet["A1"] = "Статистические данные"
    ksr_sheet["A3"] = "Число КСР"
    ksr_sheet["A4"] = "Число номеров в КСР"
    ksr_sheet["A5"] = "Число мест в КСР"
    ksr_sheet["A6"] = "Площадь номерного фонда"
    ksr_sheet["A7"] = "Численность лиц, размещенных в КСР"
    ksr_sheet["A8"] = "Число ночевок в КСР"
    ksr_sheet["A9"] = "Доходы КСР, тыс. руб."

    for year in range(2011, datetime.now().year):
        current_col = get_last_filled_column(ksr_sheet, 3) + 1
        region_row = find_cell_by_value(ksr_year_table["1"], region)[0]
        ksr_sheet.cell(row=2, column=current_col, value=year)
        ksr_sheet.cell(row=3, column=current_col, value=ksr_year_table["1"].cell(row=region_row, column=current_col + 9).value)
        ksr_sheet.cell(row=4, column=current_col, value=ksr_year_table["2"].cell(row=region_row, column=current_col + 9).value)
        ksr_sheet.cell(row=5, column=current_col, value=ksr_year_table["3"].cell(row=region_row, column=current_col + 9).value)
        ksr_sheet.cell(row=6, column=current_col, value=ksr_year_table["4"].cell(row=region_row, column=current_col + 6).value)
        ksr_sheet.cell(row=7, column=current_col, value=ksr_year_table["5"].cell(row=region_row, column=current_col + 9).value)
        ksr_sheet.cell(row=8, column=current_col, value=ksr_year_table["8"].cell(row=region_row, column=current_col + 9).value)
        ksr_sheet.cell(row=9, column=current_col, value=ksr_year_table["9"].cell(row=region_row, column=current_col + 9).value)

        if year == datetime.now().year - 1:
            ksr_sheet.cell(row=2, column=current_col + 1, value=f"{year}/2011")
            ksr_sheet.cell(row=3, column=current_col + 1, value=ksr_year_table["1"].cell(row=region_row, column=current_col + 9).value)
            ksr_sheet.cell(row=4, column=current_col + 1, value=ksr_year_table["2"].cell(row=region_row, column=current_col + 9).value)
            ksr_sheet.cell(row=5, column=current_col + 1, value=ksr_year_table["3"].cell(row=region_row, column=current_col + 9).value)
            ksr_sheet.cell(row=6, column=current_col + 1, value=ksr_year_table["4"].cell(row=region_row, column=current_col + 6).value)
            ksr_sheet.cell(row=7, column=current_col + 1, value=ksr_year_table["5"].cell(row=region_row, column=current_col + 9).value)
            ksr_sheet.cell(row=8, column=current_col + 1, value=ksr_year_table["8"].cell(row=region_row, column=current_col + 9).value)
            ksr_sheet.cell(row=9, column=current_col + 1, value=ksr_year_table["9"].cell(row=region_row, column=current_col + 9).value)

    ksr_sheet["A12"] = "Аналитические показатели"
    ksr_sheet["A13"] = f"2011 - {datetime.now().year - 1}"
    ksr_sheet["A14"] = "Темп роста КСР"
    ksr_sheet["A15"] = "Темп роста номеров"
    ksr_sheet["A16"] = "Темп роста мест"
    ksr_sheet["A17"] = "Среднее кол-во номеров на КСР"
    ksr_sheet["A18"] = "Среднее кол-во мест на КСР"
    ksr_sheet["A19"] = "Средняя площадь номера"
    ksr_sheet["A20"] = "Продолжительность проживания, ночей"
    ksr_sheet["A21"] = "Темп роста продолжительности проживания"
    ksr_sheet["A22"] = "Темп роста размещенных лиц"
    ksr_sheet["A23"] = "Темп роста ночевок"
    ksr_sheet["A24"] = "Доходы в расчете на гостя, тыс. руб."
    ksr_sheet["A25"] = "Темп роста доходов в расчете на гостя"
    ksr_sheet["A26"] = "Доходы в расчете на ночевку, тыс. руб."
    ksr_sheet["A27"] = "Темп роста доходов в расчете на ночевку"

    for year in range(2011, datetime.now().year):
        current_col = get_last_filled_column(ksr_sheet, 13) + 1
        ksr_sheet.cell(row=13, column=current_col, value=year)
        ksr_sheet.cell(
            row=14, column=current_col,
            value=f'=ROUND(({ksr_sheet.cell(row=3, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=3, column=current_col - 1).value})*100, 2)&"%"'
        )
        ksr_sheet.cell(
            row=15, column=current_col,
            value=f'=ROUND(({ksr_sheet.cell(row=4, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=4, column=current_col - 1).value})*100, 2)&"%"'
        )
        ksr_sheet.cell(
            row=16, column=current_col,
            value=f'=ROUND(({ksr_sheet.cell(row=5, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=5, column=current_col - 1).value})*100, 2)&"%"'
        )
        ksr_sheet.cell(
            row=17, column=current_col,
            value=f'={ksr_sheet.cell(row=4, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=3, column=current_col).value}'
        )
        ksr_sheet.cell(
            row=18, column=current_col,
            value=f'={ksr_sheet.cell(row=5, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=3, column=current_col).value}'
        )
        ksr_sheet.cell(
            row=19, column=current_col,
            value=f'={ksr_sheet.cell(row=6, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=4, column=current_col).value}'
        )
        ksr_sheet.cell(
            row=20, column=current_col,
            value=f'={ksr_sheet.cell(row=8, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=7, column=current_col).value}'
        )
        ksr_sheet.cell(
            row=21, column=current_col,
            value=f'=ROUND((({ksr_sheet.cell(row=20, column=current_col).value[1::]}) / '
                  f'({ksr_sheet.cell(row=20, column=current_col - 1).value[1::]}))*100, 2)&"%"'
        )
        ksr_sheet.cell(
            row=22, column=current_col,
            value=f'=ROUND(({ksr_sheet.cell(row=7, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=7, column=current_col - 1).value})*100, 2)&"%"'
        )
        ksr_sheet.cell(
            row=23, column=current_col,
            value=f'=ROUND(({ksr_sheet.cell(row=8, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=8, column=current_col - 1).value})*100, 2)&"%"'
        )
        ksr_sheet.cell(
            row=24, column=current_col,
            value=f'={ksr_sheet.cell(row=9, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=7, column=current_col).value}'
        )
        ksr_sheet.cell(
            row=25, column=current_col,
            value=f'=ROUND((({ksr_sheet.cell(row=24, column=current_col).value[1::]}) / '
                  f'({ksr_sheet.cell(row=24, column=current_col - 1).value[1::]}))*100, 2)&"%"'
        )
        ksr_sheet.cell(
            row=26, column=current_col,
            value=f'={ksr_sheet.cell(row=9, column=current_col).value} / '
                  f'{ksr_sheet.cell(row=8, column=current_col).value}'
        )
        ksr_sheet.cell(
            row=27, column=current_col,
            value=f'=ROUND((({ksr_sheet.cell(row=26, column=current_col).value[1::]}) / '
                  f'({ksr_sheet.cell(row=26, column=current_col - 1).value[1::]}))*100, 2)&"%"'
        )

    ksr_sheet.cell(row=13, column=get_last_filled_column(ksr_sheet, 13) + 1, value="Среднее")
    for row in range(14, 28):
        ksr_sheet.cell(row=row, column=16, value=get_row_average(ksr_sheet, row, 2))

    ksr_sheet["A30"] = "Месячные (оперативные) данные, статистические"
    ksr_sheet["A31"] = f"2022-{datetime.now().year}"
    ksr_sheet["B31"] = "Показатель"
    write_months_in_row(ksr_sheet, "C31")

    for year in range(2022, datetime.now().year + 1):
        current_row = get_last_filled_row(ksr_sheet, "B") + 1

        if year == 2022:
            ksr_sheet.cell(row=current_row, column=1, value=year)
            ksr_sheet.cell(row=current_row, column=2, value="Численность лиц, размещенных в КСР, нарастающим итогом")
            ksr_sheet.cell(row=current_row + 1, column=2, value="Численность лиц, размещенных в КСР, помесячно")
            ksr_sheet.cell(row=current_row + 2, column=2, value="Число ночевок в КСР, нарастающим итогом")
            ksr_sheet.cell(row=current_row + 3, column=2, value="Число ночевок в КСР, помесячно")
            ksr_sheet.cell(row=current_row + 4, column=2, value="Доходы КСР, тыс. руб., нарастающим итогом")
            ksr_sheet.cell(row=current_row + 5, column=2, value="Доходы КСР, тыс. руб., помесячно")

        else:
            ksr_sheet.cell(row=current_row, column=1, value=year)
            ksr_sheet.cell(row=current_row, column=2, value="Численность лиц, размещенных в КСР, нарастающим итогом")
            ksr_sheet.cell(row=current_row + 1, column=2, value="Численность лиц, размещенных в КСР, помесячно")
            ksr_sheet.cell(row=current_row + 2, column=2, value="Темп роста")
            ksr_sheet.cell(row=current_row + 2, column=1, value=f"{year}/{year-1}")
            ksr_sheet.cell(row=current_row + 3, column=2, value="Число ночевок в КСР, нарастающим итогом")
            ksr_sheet.cell(row=current_row + 4, column=2, value="Число ночевок в КСР, помесячно")
            ksr_sheet.cell(row=current_row + 5, column=2, value="Темп роста")
            ksr_sheet.cell(row=current_row + 5, column=1, value=f"{year}/{year - 1}")
            ksr_sheet.cell(row=current_row + 6, column=2, value="Доходы КСР, тыс. руб., нарастающим итогом")
            ksr_sheet.cell(row=current_row + 7, column=2, value="Доходы КСР, тыс. руб., помесячно")
            ksr_sheet.cell(row=current_row + 8, column=2, value="Темп роста")
            ksr_sheet.cell(row=current_row + 8, column=1, value=f"{year}/{year - 1}")

        for i in range(3, 15):
            row = get_last_filled_row(ksr_sheet, get_column_letter(i)) + 1
            region_row = find_cell_by_value(ksr_month_table[f"1.{year}"], region)[0]
            ksr_sheet.cell(row=row, column=i, value=ksr_month_table[f"1.{year}"].cell(row=region_row, column=i - 1).value)
            if i == 3:
                ksr_sheet.cell(
                    row=row + 1, column=i,
                    value=f"={ksr_sheet.cell(row=row, column=i).value}"
                )
            else:
                ksr_sheet.cell(
                    row=row + 1, column=i,
                    value=f"={ksr_sheet.cell(row=row, column=i).value} - {ksr_sheet.cell(row=row, column=i - 1).value}"
                )
            if year == 2022:
                ksr_sheet.cell(row=row + 2, column=i, value="Нет значения")
                ksr_sheet.cell(row=row + 3, column=i, value="Нет значения")
                ksr_sheet.cell(row=row + 4, column=i, value="Нет значения")
                ksr_sheet.cell(row=row + 5, column=i, value="Нет значения")
            else:
                year_row_gap = 5 if year == 2023 else 8
                ksr_sheet.cell(
                    row=row + 2, column=i,
                    value=f'=ROUND((({ksr_sheet.cell(row=row + 1, column=i).value[1::]}) / '
                          f'({str(ksr_sheet.cell(row=row - year_row_gap, column=i).value)[1::]}))*100, 2)&"%"'
                )
                ksr_sheet.cell(row=row + 3, column=i, value=ksr_month_table[f"2.{year}"].cell(row=region_row, column=i - 1).value)
                if i == 3:
                    ksr_sheet.cell(
                        row=row + 4, column=i,
                        value=f"={ksr_sheet.cell(row=row + 3, column=i).value}"
                    )
                else:
                    ksr_sheet.cell(
                        row=row + 4, column=i,
                        value=f"={ksr_sheet.cell(row=row + 3, column=i).value} - {ksr_sheet.cell(row=row + 3, column=i - 1).value}"
                    )
                ksr_sheet.cell(
                    row=row + 5, column=i,
                    value=f'=ROUND((({ksr_sheet.cell(row=row + 4, column=i).value[1::]}) / '
                          f'({str(ksr_sheet.cell(row=row + 3 - year_row_gap, column=i).value)[1::]}))*100, 2)&"%"'
                )
                ksr_sheet.cell(row=row + 6, column=i, value=ksr_month_table[f"3.{year}"].cell(row=region_row, column=i - 1).value)
                if i == 3:
                    ksr_sheet.cell(
                        row=row + 7, column=i,
                        value=f"={ksr_sheet.cell(row=row + 6, column=i).value}"
                    )
                else:
                    ksr_sheet.cell(
                        row=row + 7, column=i,
                        value=f"={ksr_sheet.cell(row=row + 6, column=i).value} - {ksr_sheet.cell(row=row + 6, column=i - 1).value}"
                    )
                ksr_sheet.cell(
                    row=row + 8, column=i,
                    value=f'=ROUND((({ksr_sheet.cell(row=row + 7, column=i).value[1::]}) / '
                          f'({str(ksr_sheet.cell(row=row - 2, column=i).value)[1::]}))*100, 2)&"%"'
                )

    current_row = get_last_filled_row(ksr_sheet, "A") + 3
    ksr_sheet.cell(row=current_row, column=1, value="Месячные (оперативные) данные, аналитические")
    ksr_sheet.cell(row=current_row + 1, column=1, value=f"2022-{datetime.now().year}")
    ksr_sheet.cell(row=current_row + 1, column=2, value="Показатель")
    write_months_in_row(ksr_sheet, f"C{current_row + 1}")

    for year in range(2023, datetime.now().year + 1):
        row = get_last_filled_row(ksr_sheet, "B") + 1
        ksr_sheet.cell(row=row, column=1, value=year)
        ksr_sheet.cell(row=row, column=2, value="Продолжительность проживания, ночей")
        ksr_sheet.cell(row=row + 1, column=2, value="Темп роста")
        ksr_sheet.cell(row=row + 1, column=1, value=f"{year}/{year - 1}")
        ksr_sheet.cell(row=row + 2, column=2, value="Доходы в расчете на гостя, тыс. руб.")
        ksr_sheet.cell(row=row + 3, column=2, value="Темп роста")
        ksr_sheet.cell(row=row + 4, column=1, value=f"{year}/{year - 1}")
        ksr_sheet.cell(row=row + 4, column=2, value="Доходы в расчете на ночевку, тыс. руб.")
        ksr_sheet.cell(row=row + 5, column=2, value="Темп роста")
        ksr_sheet.cell(row=row + 5, column=1, value=f"{year}/{year - 1}")

        for i in range(3, 15):
            base_row = 38 + (year - 2023) * 9
            ksr_sheet.cell(
                row=row, column=i,
                value=f'=ROUND((({ksr_sheet.cell(row=base_row + 4, column=i).value[1::]}) / '
                      f'({str(ksr_sheet.cell(row=base_row + 1, column=i).value)[1::]})), 1)'
            )
            ksr_sheet.cell(
                row=row + 2, column=i,
                value=f'=ROUND((({ksr_sheet.cell(row=base_row + 7, column=i).value[1::]}) / '
                      f'({str(ksr_sheet.cell(row=base_row + 1, column=i).value)[1::]})), 3)'
            )
            ksr_sheet.cell(
                row=row + 4, column=i,
                value=f'=ROUND((({ksr_sheet.cell(row=base_row + 7, column=i).value[1::]}) / '
                      f'({str(ksr_sheet.cell(row=base_row + 4, column=i).value)[1::]})), 3)'
            )
            if year == 2023:
                ksr_sheet.cell(
                    row=row + 1, column=i,
                    value="н/д"
                )
                ksr_sheet.cell(
                    row=row + 3, column=i,
                    value="н/д"
                )
                ksr_sheet.cell(
                    row=row + 5, column=i,
                    value="н/д"
                )
            else:
                ksr_sheet.cell(
                    row=row + 1, column=i,
                    value=f'=ROUND((({ksr_sheet.cell(row=row, column=i).value[1::]}) / '
                        f'({str(ksr_sheet.cell(row=row - 6, column=i).value)[1::]}))*100, 2)&"%"'
                )
                ksr_sheet.cell(
                    row=row + 3, column=i,
                    value=f'=ROUND((({ksr_sheet.cell(row=row + 2, column=i).value[1::]}) / '
                          f'({str(ksr_sheet.cell(row=row - 4, column=i).value)[1::]}))*100, 2)&"%"'
                )
                ksr_sheet.cell(
                    row=row + 5, column=i,
                    value=f'=ROUND((({ksr_sheet.cell(row=row + 4, column=i).value[1::]}) / '
                          f'({str(ksr_sheet.cell(row=row - 2, column=i).value)[1::]}))*100, 2)&"%"'
                )


    analytical_table.save(f"tables/Аналитические таблицы {region}.xlsx")


def create_all_tables(region):
    create_an_turpotok_table(region)
    create_an_mesto_no_table(region)
    create_an_ksr_table(region)
