from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string
import re
from openpyxl import load_workbook


def get_last_filled_row(work_sheet: Worksheet, column: str) -> int:
    row = work_sheet.max_row
    while row > 0 and work_sheet[f"{column}{row}"].value is None:
        row -= 1
    return row


def get_last_filled_column(work_sheet: Worksheet, row: int) -> int:
    column = work_sheet.max_column
    while column > 0 and work_sheet.cell(row=row, column=column).value is None:
        column -= 1
    return column


def write_months_in_row(ws, start_cell="A1"):
    months = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]

    col_letter = ''.join(filter(str.isalpha, start_cell))
    row = int(''.join(filter(str.isdigit, start_cell)))
    col = column_index_from_string(col_letter)

    for i, month in enumerate(months):
        ws.cell(row=row, column=col + i, value=month)


def find_cell_by_value(sheet, value):

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return [cell.row, cell.column]
    return None


def get_district_and_regions_by_region(path, sheet_name, target_region):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    current_district = None
    district_regions = {}
    found_district = None

    for row in ws.iter_rows(values_only=True):
        cell_value = str(row[0]).strip() if row[0] else ""

        if "федеральный округ" in cell_value.lower():
            current_district = cell_value
            district_regions[current_district] = []
            continue

        if current_district and cell_value:
            district_regions[current_district].append(cell_value)

            if cell_value == target_region:
                found_district = current_district

    if not found_district:
        raise ValueError(f"Регион '{target_region}' не найден в файле {path}")

    return found_district, district_regions[found_district]


def extract_numbers_from_formula(formula: str) -> list[float]:
    if not formula:
        return []
    cleaned = formula.replace(",", ".").replace("%", "")
    nums = re.findall(r"\d+\.\d+|\d+", cleaned)
    return [float(n) for n in nums]


def safe_eval_formula(formula: str) -> float | None:
    if not formula or not isinstance(formula, str):
        return None

    formula = formula.strip().lstrip("=")

    formula = (
        formula.replace("ОКРУГЛ", "round")
        .replace("ROUND", "round")
        .replace(";", ",")
        .replace(",", ".")
        .replace("&\"%\"", "")
        .replace("&'%'", "")
        .replace('"%"', "")
        .replace("%", "")
    )

    formula_clean = re.sub(r"[^0-9\.\+\-\*\/\(\),\s]", "", formula)

    try:
        result = eval(formula_clean, {"__builtins__": None, "round": round}, {})
        return float(result)
    except Exception:
        pass

    nums = re.findall(r"\d+\.\d+|\d+", formula_clean)
    if len(nums) >= 2:
        nums = [float(x) for x in nums]
        try:
            if formula.count("/") >= 3:
                a, b, c, d, *rest = nums
                return (a / b) / (c / d) * 100
            elif formula.count("/") == 1:
                a, b = nums[:2]
                return (a / b) * 100
        except Exception:
            pass
    return None


def parse_cell_value(cell_value) -> float | None:
    if cell_value is None:
        return None

    if isinstance(cell_value, (int, float)):
        return float(cell_value)

    text = str(cell_value).strip()

    if (not text or text.startswith("#") or 'кср' in text.lower() or
            'а' in text.lower()):
        return None

    if "%" in text and not text.startswith("="):
        try:
            return float(text.replace("%", "").replace(",", "."))
        except ValueError:
            return None

    if text.startswith("="):
        return safe_eval_formula(text)

    if any(op in text for op in ("/", "*", "+", "-")):
        try:
            return float(eval(text.replace(",", "."), {"__builtins__": None}, {}))
        except Exception:
            return None

    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def get_row_average(sheet: Worksheet, row: int, start_col: int = 1) -> float:
    values = []

    for col in range(start_col, sheet.max_column + 1):
        val = sheet.cell(row=row, column=col).value
        num = parse_cell_value(val)
        if num is not None:
            values.append(num)

    if not values:
        return 0.0

    return round(sum(values) / len(values), 2)
